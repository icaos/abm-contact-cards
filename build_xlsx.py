"""
Build the nametag workbook with QR codes embedded as pictures.

Produces nametag-merge.xlsx: the same columns as nametag-merge.csv, plus a
QR Code column holding the actual image, so the sheet can be checked by eye
before printing.

File:          build_xlsx.py
Author:        ICAOS
Created:       2026-09-02
Last modified: 2026-09-02

Change history:
  2026-09-02  Initial version.

TODO:
  - None outstanding.

USAGE
-----
Run from the repo root, after generate.py:

    python3 generate.py && python3 build_xlsx.py

CSV OR XLSX?
------------
A .csv is plain text and cannot contain pictures, which is why the CSV
carries a Photo File Name column instead - mail-merge software reads the
name and pulls the image from qrcodes/ at print time. That remains the
right input for a badge printer.

This workbook is for the human: it shows each person's QR code next to
their details, so a wrong or missing code is visible before 151 badges
are printed. It keeps the Photo File Name column too, so it can still
drive a mail merge if the software accepts .xlsx.

The output is gitignored (*.xlsx) - it holds real attendee details and
this repo is public.
"""

import os

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import build_csv

QRCODES_DIR = "qrcodes"
OUTPUT = "nametag-merge.xlsx"

COLUMNS = ["Counter", "First Name", "Last Name", "State", "Title",
           "Photo File Name", "QR Code"]

# Displayed size of each embedded QR, in pixels. Large enough to read and to
# scan off a screen, small enough that 151 rows stay scrollable.
QR_PX = 64
ROW_HEIGHT = 52          # points; a little over QR_PX (pixels) to leave margin
WIDTHS = [8, 16, 18, 22, 26, 26, 12]


def main():
    people = build_csv.merge_duplicates(build_csv.load_rows())

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Nametags"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C5D63")
    for column, name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column, value=name)
        cell.font = header_font
        cell.fill = header_fill
        sheet.column_dimensions[get_column_letter(column)].width = WIDTHS[column - 1]
    sheet.freeze_panes = "A2"

    missing = []
    for counter, person in enumerate(people, start=1):
        row = counter + 1
        slug = build_csv.slugify(f"{person['first']} {person['last']}")
        image_name = f"{slug}.png"
        image_path = os.path.join(QRCODES_DIR, image_name)

        values = [counter, person["first"], person["last"],
                  person["state"], person["title"], image_name]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        sheet.row_dimensions[row].height = ROW_HEIGHT

        if not os.path.exists(image_path):
            missing.append(image_name)
            sheet.cell(row=row, column=7, value="MISSING")
            continue

        picture = XLImage(image_path)
        picture.width = picture.height = QR_PX
        sheet.add_image(picture, f"G{row}")

    workbook.save(OUTPUT)
    print(f"Wrote {OUTPUT} - {len(people)} row(s), "
          f"{len(people) - len(missing)} embedded QR image(s)")

    if missing:
        print(f"\nWARNING: {len(missing)} row(s) have no image:")
        for name in missing:
            print(f"  {name}")


if __name__ == "__main__":
    main()
